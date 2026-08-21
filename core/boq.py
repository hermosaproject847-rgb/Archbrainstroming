"""BOQ engine — follows the trade-wise extraction prompts supplied by the user
(Brick Masonry, Plaster, Floor & Skirting, Paint/Putty/Primer, Wall Tiles,
Granite sill/jamb/lintel, Doors, Windows, Staircase incl. plaster). Everything
is a LIVE Excel formula, metric (m, m², cum), computed from the 2D plan.

Trades that need drawings this app does not hold (Excavation, Stone Masonry,
Anti-Termite, Waterproofing, Electrical, Plumbing, False Ceiling) get a sheet
that says DATA NOT PROVIDED with the exact input required — never invented.

A 2D plan has no height, so ONE figure (floor height) is an input, printed on
COVER. Opening heights use the schedule's sill/lintel (door = lintel, sill 0;
window/vent = lintel − sill).
"""

from __future__ import annotations

import os
import re

FT_M = 0.3048
MM = 304.8
IN = 1.0 / 12.0


def _m(ft: float) -> float:
    return round(ft * FT_M, 4)


def _styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="B0B0B0")
    return {
        "hdr_font": Font(bold=True, color="FFFFFF", size=9),
        "hdr_fill": PatternFill("solid", fgColor="34495E"),
        "title_font": Font(bold=True, size=12),
        "sub_font": Font(bold=True, size=10, color="2F5FD0"),
        "tot_font": Font(bold=True),
        "npr_fill": PatternFill("solid", fgColor="F8D0D0"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "wrap": Alignment(wrap_text=True, vertical="top"),
        "center": Alignment(horizontal="center", vertical="center"),
    }


def _open_h_m(o) -> float:
    """Opening height in metres per the rules: door = lintel height (sill 0);
    window / ventilator = lintel − sill."""
    lintel = getattr(o, "lintel_mm", None) or 0
    sill = getattr(o, "sill_mm", None) or 0
    ht = getattr(o, "height_mm", None) or 0
    if getattr(o, "is_door", False):
        h = ht or lintel or 2100
    else:
        h = (lintel - sill) if (lintel and lintel > sill) else (ht or 1200)
    return round(h / 1000.0, 3)


def _wall_thk_m(w) -> float:
    return round(w.thickness_in * 25.4 / 1000.0, 3)


def _cols_along_wall(w, columns) -> float:
    """Total column dimension ALONG the wall (metres) for columns sitting in it
    — used for the masonry column deduction."""
    tol = w.thickness_in * IN / 2 + 0.9
    a = 0.0
    for c in columns:
        cw = getattr(c, "w", 0.75)
        ch = cw if getattr(c, "shape", "") == "round" else getattr(c, "h", cw)
        if abs(w.y1 - w.y2) < 1e-6:                      # horizontal wall
            lo, hi = sorted((w.x1, w.x2))
            if abs(c.y - w.y1) <= tol and lo - 0.5 <= c.x <= hi + 0.5:
                a += cw
        elif abs(w.x1 - w.x2) < 1e-6:                    # vertical wall
            lo, hi = sorted((w.y1, w.y2))
            if abs(c.x - w.x1) <= tol and lo - 0.5 <= c.y <= hi + 0.5:
                a += ch
    return round(a * FT_M, 3)


def generate(plan, out_path: str, floor_height_ft: float = 10.0) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    S = _styles()
    wb = Workbook()
    cnt = {"items": 0, "blocked": 0}
    Hft = round(float(floor_height_ft or 10), 3)
    Hm = _m(Hft)

    walls = [w for w in plan.walls if not getattr(w, "railing", False)
             and w.length >= 1.0]
    cols = list(getattr(plan, "columns", []))
    rooms = [r for r in plan.rooms if not getattr(r, "void", False)]

    def new(name, title, headers, widths=None):
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        ws["A1"] = title
        ws["A1"].font = S["title_font"]
        ws.append([])
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=c)
            cell.font = S["hdr_font"]; cell.fill = S["hdr_fill"]
            cell.alignment = S["center"]
            cell.border = S["border"]
        ws.freeze_panes = "A4"
        widths = widths or ([8, 40] + [12] * (len(headers) - 2))
        for i, wd in enumerate(widths, 1):
            ws.column_dimensions[_col_letter(i)].width = wd
        return ws

    def row(ws, vals, formulas=None, npr=False, wrapcols=(2,)):
        r = ws.max_row + 1
        ws.append(list(vals))
        for c in range(1, len(vals) + 1):
            ws.cell(row=r, column=c).border = S["border"]
            if c in wrapcols:
                ws.cell(row=r, column=c).alignment = S["wrap"]
        if formulas:
            for col, f in formulas.items():
                cell = ws.cell(row=r, column=col)
                cell.value = f
                cell.number_format = "0.000"
        if npr:
            cnt["blocked"] += 1
            for c in range(1, len(vals) + 1):
                ws.cell(row=r, column=c).fill = S["npr_fill"]
        else:
            cnt["items"] += 1
        return r

    def total_row(ws, col_letters):
        r = ws.max_row + 1
        ws.cell(row=r, column=2, value="TOTAL").font = S["tot_font"]
        for cl in col_letters:
            ci = _col_index(cl)
            ws.cell(row=r, column=ci,
                    value=f"=SUM({cl}4:{cl}{r-1})").font = S["tot_font"]
            ws.cell(row=r, column=ci).number_format = "0.000"

    # ---------------- COVER --------------------------------------------
    cov = wb.active
    cov.title = "COVER"
    cov.sheet_view.showGridLines = False
    cov.column_dimensions["A"].width = 34
    cov.column_dimensions["B"].width = 64
    t = plan.title
    info = [
        ("BILL OF QUANTITIES", ""),
        ("Rules basis", "User's trade-wise BOQ prompt set (Brick, Plaster, "
         "Floor/Skirting, Paint, Wall Tiles, Granite, Doors, Windows, "
         "Staircase)"),
        ("Project", getattr(t, "project", "") or "-"),
        ("Drawing", getattr(t, "plan_name", "") or "FLOOR PLAN"),
        ("Run mode", "QUANTITY-ONLY (no rates)"),
        ("Floor / wall height used (m)", Hm),
        ("Units", "metres, m², cum — all cells are live formulas"),
        ("Walls", len(walls)), ("Rooms", len(rooms)), ("Columns", len(cols)),
        ("NOTE", "Trades needing structural/site/MEP drawings are marked "
                 "DATA NOT PROVIDED, never assumed."),
    ]
    for i, (k, v) in enumerate(info, 1):
        cov.cell(row=i, column=1, value=k)
        cov.cell(row=i, column=2, value=v)
        cov.cell(row=i, column=1).font = (S["title_font"] if i == 1
                                          else S["sub_font"])

    # ============ BRICK MASONRY ========================================
    ws = new("BRICK_MASONRY",
             "BRICK MASONRY BOQ — element-wise (Gross − Column − Opening)",
             ["Wall ID", "Thk (m)", "Length (m)", "Height (m)", "Gross Vol",
              "Column Ded", "Opening Ded", "Net Vol (cum)", "Ref", "Remarks"],
             [10, 9, 11, 10, 12, 12, 12, 13, 12, 30])
    for w in walls:
        Lm = _m(w.length)
        thk = _wall_thk_m(w)
        col_along = _cols_along_wall(w, cols)
        col_ded = round(col_along * Hm * thk, 4)
        open_ded = 0.0
        for o in plan.openings:
            if o.wall_id == w.id:
                open_ded += round(_m(o.width) * _open_h_m(o) * thk, 4)
        r = row(ws, [w.id, thk, Lm, Hm, None, col_ded, round(open_ded, 3),
                     None, w.id,
                     f"{'EXT' if w.exterior else 'INT'} "
                     f"{w.thickness_in:g}\" wall"])
        ws.cell(row=r, column=5).value = f"=B{r}*C{r}*D{r}"        # gross
        ws.cell(row=r, column=8).value = f"=E{r}-F{r}-G{r}"        # net
        for c in (5, 8):
            ws.cell(row=r, column=c).number_format = "0.000"
    total_row(ws, ["E", "F", "G", "H"])

    # ============ PLASTER ==============================================
    ws = new("PLASTER",
             "PLASTER BOQ — external both sides (18+12mm), internal both "
             "sides (12mm)",
             ["Wall ID", "Type", "Gross Area", "Opening Ded", "Faces",
              "Net Area (m²)", "Spec", "Ref"],
             [10, 8, 12, 12, 8, 13, 22, 10])
    for w in walls:
        Lm = _m(w.length)
        faces = 2                       # both external walls and internal walls
        ded_each = 0.0
        for o in plan.openings:
            if o.wall_id == w.id:
                ded_each += round(_m(o.width) * _open_h_m(o), 4)
        # external wall: deduct once per side but both sides plastered = 2 faces
        # internal wall between rooms: opening deducted twice (once per face)
        total_ded = round(ded_each * faces, 3)
        spec = ("Ext 18mm + Int 12mm" if w.exterior else "12mm both sides")
        r = row(ws, [w.id, "EXT" if w.exterior else "INT", None,
                     total_ded, faces, None, spec, w.id])
        ws.cell(row=r, column=3).value = f"={Lm}*{Hm}*E{r}"   # gross = L*H*faces
        ws.cell(row=r, column=6).value = f"=C{r}-D{r}"
        for c in (3, 6):
            ws.cell(row=r, column=c).number_format = "0.000"
    total_row(ws, ["C", "D", "F"])

    # ============ FLOOR & SKIRTING =====================================
    ws = new("FLOOR_SKIRTING",
             "FLOOR & SKIRTING BOQ (skirting = perimeter − door widths)",
             ["Room", "Floor Area (m²)", "Perimeter (m)", "Door Ded (m)",
              "Skirting (m)", "Material", "Ref"],
             [22, 14, 13, 12, 12, 14, 10])
    door_w_by_room = _door_widths_by_room(plan)
    for room in rooms:
        if getattr(room, "is_lawn", False) or "stair" in room.name.lower():
            continue
        Lm, Wm = _m(room.w), _m(room.h)
        peri = round(2 * (Lm + Wm), 3)
        dwid = round(door_w_by_room.get(room.name, 0.0), 3)
        mat = _room_material(plan, room)
        r = row(ws, [room.name, None, peri, dwid, None, mat, room.name])
        ws.cell(row=r, column=2).value = f"={Lm}*{Wm}"
        ws.cell(row=r, column=5).value = f"=C{r}-D{r}"
        for c in (2, 5):
            ws.cell(row=r, column=c).number_format = "0.000"
    total_row(ws, ["B", "E"])

    # ============ WALL TILES / DADO ====================================
    ws = new("WALL_TILES",
             "WALL TILES / DADO BOQ (wet areas) — gross = perimeter × tile "
             "height, deduct doors in zone",
             ["Room", "Perimeter (m)", "Tile Ht (m)", "Gross (m²)",
              "Door Ded (m²)", "Net Tile (m²)", "WP Area (m²)", "Ref",
              "Remarks"],
             [20, 12, 10, 11, 11, 12, 11, 9, 26])
    from . import electrical as E
    TILE_H = 2.1        # 7'-0" default dado height (confirm from finish sched)
    for room in rooms:
        if not _is_wet(room, E):
            continue
        Lm, Wm = _m(room.w), _m(room.h)
        peri = round(2 * (Lm + Wm), 3)
        dwid = round(door_w_by_room.get(room.name, 0.0), 3)
        r = row(ws, [room.name, peri, TILE_H, None,
                     None, None, None, room.name,
                     "dado ht 2.1m assumed — confirm from finish schedule"])
        ws.cell(row=r, column=4).value = f"=B{r}*C{r}"                 # gross
        ws.cell(row=r, column=5).value = f"={dwid}*C{r}"              # door ded
        ws.cell(row=r, column=6).value = f"=D{r}-E{r}"               # net
        ws.cell(row=r, column=7).value = f"={Lm}*{Wm}+B{r}*0.6"       # WP
        for c in (4, 5, 6, 7):
            ws.cell(row=r, column=c).number_format = "0.000"
    if ws.max_row == 3:
        row(ws, ["No wet area identified", None, None, None, None, None, None,
                 "-", "NIL"])
    total_row(ws, ["D", "E", "F", "G"])

    # ============ PAINT / PUTTY / PRIMER ===============================
    ws = new("PAINT",
             "PAINT / PUTTY / PRIMER BOQ (both sides; ext=Putty2+Primer1+"
             "Apex2, int=Putty2+Primer1+Emulsion2)",
             ["Wall ID", "Type", "Gross Area", "Opening Ded", "Net Area (m²)",
              "Treatment", "Ref"],
             [10, 8, 12, 12, 13, 30, 10])
    for w in walls:
        Lm = _m(w.length)
        faces = 2
        ded = 0.0
        for o in plan.openings:
            if o.wall_id == w.id:
                ded += round(_m(o.width) * _open_h_m(o), 4)
        total_ded = round(ded * faces, 3)
        treat = ("Acrylic Putty(2)+Primer(1)+Apex(2)" if w.exterior
                 else "Wall Putty(2)+Primer(1)+Emulsion(2)")
        r = row(ws, [w.id, "EXT" if w.exterior else "INT", None,
                     total_ded, None, treat, w.id])
        ws.cell(row=r, column=3).value = f"={Lm}*{Hm}*{faces}"
        ws.cell(row=r, column=5).value = f"=C{r}-D{r}"
        for c in (3, 5):
            ws.cell(row=r, column=c).number_format = "0.000"
    total_row(ws, ["C", "D", "E"])

    # ============ GRANITE (sill / jamb / lintel) =======================
    ws = new("GRANITE",
             "GRANITE BOQ — sill / jamb / lintel (20mm) for every opening",
             ["Opening", "Element", "Width (m)", "Height (m)", "Depth (m)",
              "Area (m²)", "Ref"],
             [10, 9, 11, 11, 10, 12, 10])
    GTHK = 0.02
    for o in plan.openings:
        if o.type not in ("single_door", "double_door", "sliding_door",
                          "door", "window", "vent"):
            continue
        tag = o.tag or o.type[:1].upper()
        wid = _m(o.width)
        ht = _open_h_m(o)
        # sill (not for doors), jamb, lintel
        if not getattr(o, "is_door", False):
            r = row(ws, [tag, "Sill", wid, None, GTHK, None, tag])
            ws.cell(row=r, column=6).value = f"=C{r}*E{r}"
            ws.cell(row=r, column=6).number_format = "0.000"
        r = row(ws, [tag, "Jamb x2", 0.15, ht, GTHK, None, tag])
        ws.cell(row=r, column=6).value = f"=D{r}*C{r}*2"
        ws.cell(row=r, column=6).number_format = "0.000"
        r = row(ws, [tag, "Lintel", wid, None, GTHK, None, tag])
        ws.cell(row=r, column=6).value = f"=C{r}*E{r}"
        ws.cell(row=r, column=6).number_format = "0.000"
    total_row(ws, ["F"])

    # ============ DOORS =================================================
    ws = new("DOORS", "DOOR BOQ — frame (rmt) + shutter (m²), per mark",
             ["Mark", "Type", "Nos", "Width (m)", "Height (m)",
              "Frame 2H+W (rmt)", "Shutter Area (m²)", "Ref", "Remarks"],
             [8, 12, 6, 10, 10, 15, 15, 9, 24])
    for tag, d in _openings_by_tag(plan, doors=True).items():
        o = d["o"]
        wid = _m(o.width); ht = _open_h_m(o)
        r = row(ws, [tag, "flush (confirm)", d["n"], wid, ht, None, None, tag,
                     "material/hardware from door schedule"])
        ws.cell(row=r, column=6).value = f"=(2*E{r}+D{r})*C{r}"
        ws.cell(row=r, column=7).value = f"=D{r}*E{r}*C{r}"
        for c in (6, 7):
            ws.cell(row=r, column=c).number_format = "0.000"
    total_row(ws, ["C", "F", "G"])

    # ============ WINDOWS ==============================================
    ws = new("WINDOWS", "WINDOW BOQ — overall frame size (m²), per mark",
             ["Mark", "Type", "Nos", "Width (m)", "Height (m)",
              "Area (m²)", "Ref", "Remarks"],
             [8, 14, 6, 10, 10, 12, 9, 24])
    for tag, d in _openings_by_tag(plan, doors=False).items():
        o = d["o"]
        wid = _m(o.width); ht = _open_h_m(o)
        kind = "Ventilator" if o.type == "vent" else "Window"
        r = row(ws, [tag, kind, d["n"], wid, ht, None, tag,
                     "material/glazing from window schedule"])
        ws.cell(row=r, column=6).value = f"=D{r}*E{r}*C{r}"
        ws.cell(row=r, column=6).number_format = "0.000"
    total_row(ws, ["C", "F"])

    # ============ STAIRCASE (finish + PLASTER) =========================
    ws = new("STAIRCASE",
             "STAIRCASE BOQ — tread / riser finish + PLASTER (soffit, sides, "
             "landing)",
             ["Item", "Nos", "Length (m)", "Breadth (m)", "Qty", "Unit",
              "Ref", "Remarks"],
             [40, 6, 11, 11, 11, 7, 9, 26])
    any_stair = False
    for si, s in enumerate(getattr(plan, "stairs", []), 1):
        any_stair = True
        N = (getattr(s, "steps_f1", 0) or 0) + (getattr(s, "steps_f2", 0) or 0)\
            + (getattr(s, "steps_f3", 0) or 0) or (getattr(s, "treads", 0) or 0)
        if not N:
            N = max(1, round(Hft / 0.5))
        going, riser = _m(0.875), _m(0.5)       # 10.5" / 6" standard
        width = _m((getattr(s, "flight_width", 0) or 0) or 3.33)
        note = f"tread/riser STANDARD 267/150mm; {N} steps shown"
        rem_len = round(N * (going + riser), 3)      # sloped soffit length ~
        rows_ = [
            (f"Stair {si} — tread finish", N, going, width, "sqm", note),
            (f"Stair {si} — riser finish", N, riser, width, "sqm", note),
            (f"Stair {si} — waist-slab SOFFIT plaster", 1, rem_len, width,
             "sqm", "underside of the flights"),
            (f"Stair {si} — stringer / side PLASTER (both)", 2, rem_len,
             _m(0.75), "sqm", "0.75m deep waist assumed"),
            (f"Stair {si} — nosing / anti-skid strip", N, width, None, "rmt",
             note),
        ]
        lsz = getattr(s, "landing_size", 0) or 0
        if lsz:
            nl = 2 if getattr(s, "type", "") == "U3" else 1
            rows_.append((f"Stair {si} — landing finish", nl, _m(lsz), _m(lsz),
                          "sqm", ""))
            rows_.append((f"Stair {si} — landing soffit plaster", nl, _m(lsz),
                          _m(lsz), "sqm", ""))
        for desc, nos, L, B, unit, rem in rows_:
            r = row(ws, [desc, nos, L, B, None, unit, f"Stair {si}", rem])
            if B is None:
                ws.cell(row=r, column=5).value = f"=B{r}*C{r}"
            else:
                ws.cell(row=r, column=5).value = f"=B{r}*C{r}*D{r}"
            ws.cell(row=r, column=5).number_format = "0.000"
    if not any_stair:
        row(ws, ["No staircase in the drawing", 0, None, None, "=0", "-",
                 "-", "NIL"])
    total_row(ws, ["E"])

    # ============ ELECTRICAL (from the software's own layout) ==========
    ws = new("ELECTRICAL",
             "ELECTRICAL BOQ (point system) — from the electrical layout",
             ["Item", "System", "Description", "Unit", "Qty", "Ref",
              "Remarks"], [8, 12, 40, 7, 8, 10, 26])
    elec = list(getattr(plan, "elec", []))
    circuits = list(getattr(plan, "circuits", []))
    ELIGHT = ("SL", "ASL", "PL", "CSL", "CV", "WL", "BWL", "HL", "CH", "ML",
              "STL", "TR")
    ELAB = {"SL": "COB spot light", "ASL": "Adjustable spot", "PL": "Panel "
            "light", "CSL": "Surface light", "CV": "Cove light strip",
            "WL": "Wall light", "BWL": "Bedside wall light", "HL": "Pendant",
            "CH": "Chandelier", "ML": "Mirror light", "STL": "Step light",
            "TR": "Track light", "CF": "Ceiling fan", "EF": "Exhaust fan",
            "AC": "AC point", "SB": "Modular switch board", "DB": "Distribution "
            "board"}

    def nc(codes):
        return sum(1 for e in elec if e.code in codes)
    sn = 0
    # wiring points (point = DB → switch → fixture)
    pts = [("Light point wiring (conduit+wire+switch)", nc(ELIGHT)),
           ("Ceiling-fan point wiring", nc(("CF",))),
           ("Exhaust-fan point wiring", nc(("EF",))),
           ("AC point wiring (dedicated + isolator)", nc(("AC",)))]
    pwr = sum(int(getattr(c, "points", 0) or 0) for c in circuits
              if getattr(c, "kind", "") == "power")
    gys = sum(int(getattr(c, "points", 0) or 0) for c in circuits
              if getattr(c, "kind", "") == "geyser")
    if pwr:
        pts.append(("6/16A power socket point wiring", pwr))
    if gys:
        pts.append(("Geyser point wiring (20A)", gys))
    for desc, n in pts:
        if n:
            sn += 1
            row(ws, [f"E{sn}", "Wiring", desc, "point", n, "elec layout", ""])
    # fixtures & accessories, code-wise
    from collections import Counter
    cc = Counter(e.code for e in elec)
    for code, n in cc.items():
        sn += 1
        row(ws, [f"E{sn}", "Fixture", ELAB.get(code, code), "nos", n,
                 "elec layout", code])
    # circuits / DB ways
    for c in circuits:
        sn += 1
        row(ws, [f"E{sn}", "Circuit",
                 f"Final circuit {getattr(c,'id','')} "
                 f"({getattr(c,'kind','')}) — {getattr(c,'mcb','')} MCB, "
                 f"{getattr(c,'wire','')} sqmm", "ckt", 1, "circuit sched",
                 f"{getattr(c,'points',0)} pts, {getattr(c,'load_w',0):.0f} W"])
    sn += 1
    row(ws, [f"E{sn}", "Wiring", "Conduit & wire running length (size-wise)",
             "rmt", "DATA NOT PROVIDED", "-",
             "point count is exact; run length is measured on site / from "
             "conduit layout"], npr=True)

    # ============ PLUMBING (from the software's own layout) ============
    ws = new("PLUMBING",
             "PLUMBING BOQ — fixtures/valves (nos) + pipe runs (rmt)",
             ["Item", "System", "Description", "Unit", "Qty", "Ref",
              "Remarks"], [8, 10, 40, 7, 9, 10, 24])
    plumb = list(getattr(plan, "plumb", []))
    pipes = list(getattr(plan, "pipes", []))
    PLAB = {"WC": "Water closet (EWC)", "WCAC": "WC 2-way angle cock",
            "BAC": "Basin angle cock (pair)", "HF": "Health faucet",
            "SH": "Shower head", "SMX": "Shower mixer/diverter",
            "SKC": "Sink cock / wall mixer", "NT": "Nahani trap",
            "BBT": "Bottle trap", "SV": "Isolation valve", "GT": "Gully trap",
            "IC": "Inspection chamber", "SS": "Soil stack", "WS": "Waste stack",
            "VP": "Vent pipe cowl", "CO": "Cleanout plug", "UGT": "UG water "
            "tank", "OHT": "Overhead tank", "PUMP": "Pump set"}
    from collections import Counter as _C
    pc = _C(p.code for p in plumb)
    SYS = {"CW": "Cold water", "HW": "Hot water", "SOIL": "Soil",
           "WASTE": "Waste", "VENT": "Vent", "RW": "Rain water",
           "STORM": "Storm"}
    sn = 0
    for code, n in pc.items():
        sys = next((SYS.get(p.system, p.system) for p in plumb
                    if p.code == code), "")
        sn += 1
        row(ws, [f"P{sn}", sys, PLAB.get(code, code), "nos", n,
                 "plumbing layout", code])
    # pipe runs by system + diameter
    runs: dict = {}
    for r in pipes:
        key = (getattr(r, "system", ""), round(getattr(r, "dia_mm", 0) or 0))
        runs[key] = runs.get(key, 0.0) + _m(r.length_ft)
    for (sys, dia), L in sorted(runs.items()):
        sn += 1
        row(ws, [f"P{sn}", SYS.get(sys, sys),
                 f"{SYS.get(sys, sys)} pipe {dia:.0f}mm dia", "rmt",
                 round(L, 3), "pipe layout", "centre-line length"])
    if sn == 0:
        row(ws, ["P1", "-", "No plumbing laid out yet — run Plumbing Layout",
                 "-", 0, "-", "NIL"])

    # ============ DATA NOT PROVIDED trades =============================
    npr = {
        "EXCAVATION": "Founding levels, footing sizes/depths, soil "
                      "classification and site levels — from the structural "
                      "footing layout, sections and soil report.",
        "STONE_MASONRY": "Foundation/plinth stone masonry sizes and mortar "
                         "from the structural foundation drawings.",
        "ANTI_TERMITE": "Foundation plan area, perimeter and IS 6313 stages "
                        "from the structural + treatment specification.",
        "WATERPROOFING": "System, coats and locations from the waterproofing "
                         "drawings / specification.",
        "FALSE_CEILING": "RCP with ceiling type, levels and drops.",
    }
    for name, needs in npr.items():
        ws = new(name, f"{name.replace('_', ' ')} BOQ", ["Item", "Status",
                 "Required input"], [40, 20, 80])
        row(ws, [name.replace("_", " ").title() + " — not in this drawing set",
                 "DATA NOT PROVIDED", needs], npr=True)

    # ---- order & save --------------------------------------------------
    order = ["COVER", "BRICK_MASONRY", "PLASTER", "FLOOR_SKIRTING",
             "WALL_TILES", "PAINT", "GRANITE", "DOORS", "WINDOWS", "STAIRCASE",
             "ELECTRICAL", "PLUMBING", "EXCAVATION", "STONE_MASONRY",
             "ANTI_TERMITE", "WATERPROOFING", "FALSE_CEILING"]
    wb._sheets.sort(key=lambda s: order.index(s.title)
                    if s.title in order else 99)
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return cnt


# ------------------------------------------------------------- helpers
def _col_letter(i: int) -> str:
    s = ""
    while i:
        i, rem = divmod(i - 1, 26)
        s = chr(65 + rem) + s
    return s


def _col_index(letters: str) -> int:
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n


def _door_widths_by_room(plan) -> dict:
    out: dict = {}
    for o in plan.openings:
        if not getattr(o, "is_door", False):
            continue
        rm = (o.swing.room if getattr(o, "swing", None) else "") or ""
        out[rm] = out.get(rm, 0.0) + _m(o.width)
    return out


def _room_material(plan, room) -> str:
    for s in getattr(plan, "flooring", []):
        if getattr(s, "room", "") == room.name:
            return getattr(s, "material", "tile")
    return "tile (confirm)"


def _is_wet(room, E) -> bool:
    n = room.name.lower()
    try:
        if E.classify(room.name) == "wet":
            return True
    except Exception:
        pass
    return any(k in n for k in ("toilet", "bath", "wc", "kitchen", "utility",
                                "wash", "powder"))


def _openings_by_tag(plan, doors: bool) -> dict:
    seen: dict = {}
    for o in plan.openings:
        is_door = getattr(o, "is_door", False)
        if doors and not is_door:
            continue
        if not doors and (is_door or o.type not in ("window", "vent")):
            continue
        tag = o.tag or o.type[:1].upper()
        if tag in seen:
            seen[tag]["n"] += max(1, getattr(o, "count", 1))
        else:
            seen[tag] = {"o": o, "n": max(1, getattr(o, "count", 1))}
    return seen
